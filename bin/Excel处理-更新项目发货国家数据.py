import os
import json
from openpyxl import load_workbook

# -*- coding: utf-8 -*-

def read_xlsx(excel_file):
    workbook = load_workbook(excel_file, data_only=True)
    sheet = workbook['Sheet1']
    # 读取数据时过滤掉包含空值的行
    data = [list(row) for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None and row[1] is not None]
    return data

def write_json_to_file(json_data, file_path):
    with open(file_path, 'w', encoding='utf-8') as file:
        json.dump(json_data, file, indent=2, ensure_ascii=False)

if __name__ == "__main__":
    script_directory = os.path.dirname(os.path.abspath(__file__))
    parent_directory = os.path.dirname(script_directory)

    excel_file = os.path.join(script_directory, "shipping_country", "商城发货国家001.xlsx")

    output_path = os.path.join(parent_directory, "src", "language", "country", "en.json")

    print(f'开始读取Excel: {excel_file}')

    sheet1_content = read_xlsx(excel_file)
    print('> 开始排序...')
    # sheet1_content.sort(key=lambda x: x[1])  # 按第二个元素（value）排序

    print('> 格式化...')
    json_data = {"list": []}
    for item in sheet1_content:
        json_data["list"].append({"value": item[1], "label": item[2]})

    print('获取数据长度:',json_data['list'].__len__())

    print('> 写入...')
    write_json_to_file(json_data, output_path)

    print('更新完成✅')