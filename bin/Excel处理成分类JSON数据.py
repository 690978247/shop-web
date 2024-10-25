import os
import json
from openpyxl import load_workbook
from collections import OrderedDict
import re
# 请确保在文件的开头添加这一行，以支持中文字符
# -*- coding: utf-8 -*-

def capitalize_after_special_chars(string):
    # 定义一个正则表达式，用于匹配特殊字符后的首个字母
    pattern = re.compile(r"([&/ ])(\w)")
    
    # 使用正则表达式的sub方法进行匹配和替换
    # 这里使用了一个lambda表达式，它会将匹配到的组（g(1)是特殊字符，g(2)是首个字母）进行处理，保持特殊字符不变，并将首字母大写
    result = pattern.sub(lambda m: m.group(1) + m.group(2).upper(), string)
    
    return result


#value字符串处理
def  capitalize_string(string):
    #空格后面大写
    words = string.split(' ')
    # 对列表中的每个单词首字母进行大写处理
    capitalized_words = [word[0].upper() + word[1:] if word else '' for word in words]
    # 将处理后的单词列表重新连接成字符串
    capitalized_string = ' '.join(capitalized_words)
    #对&、/、'后的首字母大写
    target = capitalize_after_special_chars(capitalized_string)
    print(f"{words} => {target}")
    return target


def read_xlsx(excel_file):
    workbook = load_workbook(excel_file)
    result = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = []

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            data.append(list(row))

        result[sheet_name] = data

    return result

def get_final_json(final_data_map, language):
    result = {
        "list": extract_language_data(final_data_map, language)
    }
    return json.dumps(result, indent=2, ensure_ascii=False)

def merge_list_map(data_list):
    merged_data = {}

    for item in data_list:
        category_id = item["categoryID"]

        if category_id in merged_data:
            parent_item = merged_data[category_id]
            parent_item["children"].extend(item["children"])
        else:
            merged_item = {
                "name": item["name"],
                "pid": item["pid"],
                "categoryID": item["categoryID"],
                "name_en": item["name_en"],
                "name_pt": item["name_pt"],
                "name_es": item["name_es"],
                "children": item["children"]
            }
            merged_data[category_id] = merged_item

    merged_list = list(merged_data.values())
    return merged_list

def merge_all_list_map(data_list_1, data_list_2):
    merged_data = OrderedDict()

    for index, item in enumerate(data_list_1):
        category_id = item["categoryID"]

        if category_id in merged_data:
            parent_item = merged_data[category_id]
            parent_item["children"].extend(create_children_list(item))
            parent_item["children"].sort(key=lambda x: len(str(x)), reverse=True)
        else:
            children_list = create_children_list(item)
            merged_item = {
                **item,
                "children": sorted(children_list, key=lambda x: len(str(x)), reverse=True)
            }
            merged_data[category_id] = merged_item

    for item in data_list_2:
        pid = item["pid"]

        if pid in merged_data:
            parent_item = merged_data[pid]
            parent_item["children"].append(item)
            parent_item["children"].sort(key=lambda x: len(str(x)), reverse=True)
        else:
            # 假设item["children"]已存在并需要被排序
            children_list = sorted(item.get("children", []), key=lambda x: len(str(x)), reverse=True)
            merged_item = {
                **item,
                "children": children_list
            }
            merged_data[pid] = merged_item

    # 保持输入顺序
    merged_list = [merged_data[category_id] for category_id in (item["categoryID"] for item in data_list_1) if category_id in merged_data]

    return merged_list

def create_children_list(item):
    children_list = item.get("children", [])
    return children_list

def extract_language_data(data_list, language):
    extracted_data = []

    for item in data_list:
        if "icon" in item:
            extracted_item = {
                "icon": item["icon"],
                "name": item["name"] if language == 'zh' else item[f"name_{language}"],
                "categoryID": item["categoryID"]
            }
        else:
            extracted_item = {
                "name": item["name"] if language == 'zh' else item[f"name_{language}"],
                "categoryID": item["categoryID"]
            }

        if "children" in item:
            extracted_children = extract_language_data(item["children"], language)
            extracted_item["children"] = extracted_children

        extracted_data.append(extracted_item)

    return extracted_data

def write_json_to_file(json_string, file_path):
    with open(file_path, 'w', encoding='utf-8') as file:  # 使用 'utf-8' 编码写入文件
            file.write(json_string)


if __name__ == "__main__":

    # 获取当前脚本所在目录
    script_directory = os.path.dirname(os.path.abspath(__file__))

    # 获取脚本的上级目录
    parent_directory = os.path.dirname(script_directory)

    excel_file = os.path.join(script_directory, "category", "一级类目和二级类目葡语和西语翻译.xlsx")
    output_directory = parent_directory + "/src/language/category"

    print(f'开始读取Excel: {excel_file}')
    result = read_xlsx(excel_file)
    level1_result = result['一级类目']
    level2_result = result['二级类目']
    print(f'读取Excel结束')

    print(f'开始转换为JSON')
    level1 = []
    for i, model in enumerate(level1_result):
        item = {
            "icon": f"icon-a-{i + 1}",
            "name": model[0],
            "categoryID": model[1],
            "name_en": capitalize_string(model[2]),
            "name_pt": capitalize_string(model[3]),
            "name_es":capitalize_string(model[4]),
            "children": []
        }
        level1.append(item)

    level2 = []
    for i, model in enumerate(level2_result):
        children = [{
            "name": model[6],
            "categoryID": model[7],
            "name_en": capitalize_string(model[8]),
            "name_pt": capitalize_string(model[9]),
            "name_es": capitalize_string(model[10]),
        }]
        item = {
            "pid": model[0],
            "name": model[1],
            "name_en": capitalize_string(model[2]),
            "name_pt": capitalize_string(model[3]),
            "name_es": capitalize_string(model[4]),
            "categoryID": model[5],
            "children": children,
        }
        level2.append(item)

    merged_list_map = merge_list_map(level2)
    final_data_map = merge_all_list_map(level1, merged_list_map)

    out_string_map = {
        'zh.json': get_final_json(final_data_map, 'zh'),
        'en.json': get_final_json(final_data_map, 'en'),
        'es.json': get_final_json(final_data_map, 'es'),
        'br.json': get_final_json(final_data_map, 'pt')
    }

    for key, value in out_string_map.items():
        write_json_to_file(value, f"{output_directory}/{key}")

    print(f'全部数据处理完成: {output_directory}')
    